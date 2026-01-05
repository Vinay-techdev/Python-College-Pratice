
import openpyxl

wb = openpyxl.load_workbook('Sample_Sales_Data.xlsx')
print(type(wb))

#? call the worksheet names
ws = wb['SalesData']
print(ws)

#? count the number of rows and columns in the worksheet
rows = ws.max_row
print("Number of rows: ", rows)

columns = ws.max_column
print("Number of rows: ", columns)

#? Read the data from the Excel file
value = ws['A1'].value
print("\nReading Excel\n",value)
value = ws['B2'].value
print(value)

#? Reading mutliple cells from the Excel file
a = ws['A1'].value
b = ws['B1'].value
c = ws['C1'].value

print("\nReading multiple value\n",a, b, c)

#? Reading the full row
for cell in ws[2]:
    print(cell.value)

for cell in ws[3]:
    print(cell.value)

#? Reading the full column
for cell in ws['A']:
    print(cell.value)

#? Writing to a cell using cell address
ws['D1'] = 'Total'

from openpyxl import load_workbook
wb = load_workbook('Sample_Sales_Data.xlsx')
ws = wb['SalesData']
ws.cell(row = 1, column= 4, value= 'Total')

for i, row in enumerate(ws.iter_rows(min_row= 2, max_col= 3), start= 2):
    value = row[2].value
    ws.cell(row= i, column= 4, value= value)

wb.save('Sample_Sales_Data.xlsx') 

#? Adding new rows into worksheet
from openpyxl import load_workbook
wb = load_workbook('Sample_Sales_Data.xlsx')
ws = wb.active
ws.append(['Name', 'Age', 'Score'])
ws.append(['Alice', 25, 90])
ws.append(['Bob', 23, 80])
wb.save('Sample_Sales_Data.xlsx')